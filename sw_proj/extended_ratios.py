#*****************************************************************************************************#
# PROJECT:
# Facial Landmark Detection and Ratio Analysis
#*****************************************************************************************************# 

#*****************************************************************************************************#
# IMPORT AREA
#*****************************************************************************************************# 
import os
import cv2
import dlib
import numpy as np
import pandas as pd
import openpyxl
import matplotlib.pyplot as plt
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

#*****************************************************************************************************#
# CONFIGURATION
#*****************************************************************************************************# 
IMAGE_FOLDER = "sw_proj/Wanted_DeadOrAlive"
OUTPUT_IMG_DIR = "sw_proj/face_with_points"
OUTPUT_EXCEL = "face_landmarks_extended.xlsx"
TABLE_EXCEL = True

#*****************************************************************************************************#
# CLASS DEFINITION
#*****************************************************************************************************#
class FaceLandmarkProcessor:
    def __init__(self, predictor_path):
        self.detector = dlib.get_frontal_face_detector()
        self.predictor = dlib.shape_predictor(predictor_path)
        self.landmark_data = []
        self.ratios = []
        self.extended_ratios = []

        os.makedirs(OUTPUT_IMG_DIR, exist_ok=True)

    def calculate_distance(self, p1, p2):
        return np.linalg.norm(np.array(p1) - np.array(p2))

    def calculate_r_value(self, landmarks):
        height = self.calculate_distance(landmarks[27], landmarks[8])
        width = self.calculate_distance(landmarks[0], landmarks[16])
        return height / width if width else 0

    def calculate_all_ratios(self, landmarks):
        d = self.calculate_distance
        r = {}
        r["r1_height/width"] = d(landmarks[27], landmarks[8]) / d(landmarks[0], landmarks[16]) if d(landmarks[0], landmarks[16]) else 0
        r["r2_faceHeight/noseHeight"] = d(landmarks[27], landmarks[8]) / d(landmarks[27], landmarks[33]) if d(landmarks[27], landmarks[33]) else 0
        r["r3_faceWidth/noseWidth"] = d(landmarks[0], landmarks[16]) / d(landmarks[31], landmarks[35]) if d(landmarks[31], landmarks[35]) else 0
        left_eye = d(landmarks[36], landmarks[39])
        right_eye = d(landmarks[42], landmarks[45])
        eye_dist = d(landmarks[39], landmarks[42])
        mean_eye_width = (left_eye + right_eye) / 2
        r["r4_eyeWidth/eyeDistance"] = mean_eye_width / eye_dist if eye_dist else 0
        r["r5_faceWidth/eyeWidth"] = d(landmarks[0], landmarks[16]) / mean_eye_width if mean_eye_width else 0
        r["r6_faceWidth/eyeDistance"] = d(landmarks[0], landmarks[16]) / eye_dist if eye_dist else 0
        r["r7_mouthWidth/mouthHeight"] = d(landmarks[48], landmarks[54]) / d(landmarks[51], landmarks[57]) if d(landmarks[51], landmarks[57]) else 0
        r["r8_faceWidth/mouthWidth"] = d(landmarks[0], landmarks[16]) / d(landmarks[48], landmarks[54]) if d(landmarks[48], landmarks[54]) else 0
        r["r9_faceHeight/mouthHeight"] = d(landmarks[27], landmarks[8]) / d(landmarks[51], landmarks[57]) if d(landmarks[51], landmarks[57]) else 0
        diff = abs(d(landmarks[36], landmarks[27]) - d(landmarks[27], landmarks[45]))
        dist_36_45 = d(landmarks[36], landmarks[45])
        r["r10_asymmetry_diff"] = diff
        r["r10_ratio"] = dist_36_45 / diff if diff else 0
        r["r10_half_ratio"] = (dist_36_45 / 2 - d(landmarks[36], landmarks[27]))
        r["r10_asymmetry_percent"] = ((r["r10_half_ratio"]) / (dist_36_45 / 2)) * 100 if dist_36_45 else 0
        return r

    def calculate_dispersion(self, ratios, mean_value):
        if isinstance(ratios, (list, np.ndarray)):  # Verificăm dacă ratios este o listă sau un array
            dispersion = np.sqrt(np.sum((np.array(ratios) - mean_value) ** 2) / (len(ratios) - 1))
        else: 
            dispersion = 0
        return dispersion

    def draw_landmarks(self, image, landmarks, filename):
        for pt in landmarks:
            cv2.circle(image, pt, 2, (0, 255, 0), -1)
        cv2.line(image, landmarks[27], landmarks[8], (0, 0, 255), 2)
        cv2.line(image, landmarks[0], landmarks[16], (0, 0, 255), 2)
        path = os.path.join(OUTPUT_IMG_DIR, filename)
        cv2.imwrite(path, image)

    def process_images(self, folder):
        for file in os.listdir(folder):
            if file.lower().endswith((".jpg", ".jpeg", ".png")):
                path = os.path.join(folder, file)
                image = cv2.imread(path)
                gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
                faces = self.detector(gray)
                for face in faces:
                    landmarks_obj = self.predictor(gray, face)
                    landmarks = [(landmarks_obj.part(i).x, landmarks_obj.part(i).y) for i in range(68)]
                    self.draw_landmarks(image.copy(), landmarks, file)
                    row = {"image": file}
                    for i, (x, y) in enumerate(landmarks):
                        row[f"point_{i}_x"] = x
                        row[f"point_{i}_y"] = y
                    self.landmark_data.append(row)
                    self.ratios.append({"image": file, "R": self.calculate_r_value(landmarks)})
                    r_all = self.calculate_all_ratios(landmarks)
                    r_all["image"] = file
                    self.extended_ratios.append(r_all)

    def save_to_excel(self):
        with pd.ExcelWriter(OUTPUT_EXCEL, engine='openpyxl') as writer:
            # Crearea și salvarea sheet-ului "Overview"
            df_landmarks = pd.DataFrame(self.landmark_data)
            df_landmarks.to_excel(writer, sheet_name="Overview", index=False)

            if TABLE_EXCEL:
                wb = writer.book
                ws = wb["Overview"]
                table = Table(displayName="FaceLandmarkTable", ref=f"A1:{get_column_letter(len(df_landmarks.columns))}{len(df_landmarks)+1}")
                style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
                table.tableStyleInfo = style
                ws.add_table(table)

            pd.DataFrame(self.ratios).to_excel(writer, sheet_name="R_values", index=False)

            stats = []
            for ratio in self.ratios:
                image_name = ratio["image"]
                row_stats = {"image": image_name}
                all_ratios = {key: [] for key in ratio if key != "image"}

                for key in ratio:
                    if key != "image":  
                        r_values = ratio[key]
                        all_ratios[key].append(r_values)
                        
                for key, values in all_ratios.items():
                    mean_value = np.mean(values)
                    dispersion = np.sqrt(np.sum((np.array(values) - mean_value) ** 2) / (len(values) - 1))
                    row_stats[f"{key}_Mean"] = mean_value
                    row_stats[f"{key}_Dispersion"] = dispersion

                stats.append(row_stats)

            df_stats = pd.DataFrame(stats)

            df_stats.to_excel(writer, sheet_name="Stats", index=False)

            if TABLE_EXCEL:
                wb = writer.book
                ws = wb["Stats"]
                table = Table(displayName="StatsTable", ref=f"A1:{get_column_letter(len(df_stats.columns))}{len(df_stats)+1}")
                style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
                table.tableStyleInfo = style
                ws.add_table(table)

            df_ext = pd.DataFrame(self.extended_ratios)
            df_ext.to_excel(writer, sheet_name="Extended_Ratios", index=False)

            if TABLE_EXCEL:
                wb = writer.book
                ws = wb["Extended_Ratios"]
                table = Table(displayName="ExtendedRatiosTable", ref=f"A1:{get_column_letter(len(df_ext.columns))}{len(df_ext)+1}")
                style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
                table.tableStyleInfo = style
                ws.add_table(table)

if __name__ == "__main__":
    predictor_model = "sw_proj/shape_predictor_68_face_landmarks.dat"  
    processor = FaceLandmarkProcessor(predictor_model)
    processor.process_images(IMAGE_FOLDER)
    processor.save_to_excel()
    print(f"Procesare completă. Fișierul salvat: {OUTPUT_EXCEL}")
