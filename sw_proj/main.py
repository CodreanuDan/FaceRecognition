#*****************************************************************************************************#
# PROJECT:
#*****************************************************************************************************# 
'''
    Some useful information
    https://medium.com/analytics-vidhya/building-an-api-with-dlib-python-heroku-2d25687e66f0
'''
#*****************************************************************************************************# 


#*****************************************************************************************************#
# IMPORT AREA
#*****************************************************************************************************# 
import os
import cv2
import dlib
import numpy as np
import pandas as pd
import openpyxl
import matplotlib.pyplot as plt
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.worksheet.table import Table, TableStyleInfo



#*****************************************************************************************************#
# IMPLEMENATION
#*****************************************************************************************************# 

# =========================
# CONFIGURATION FLAGS
# =========================
TABLE_EXCEL = True
IMAGE_FOLDER = "sw_proj/Wanted_DeadOrAlive"
OUTPUT_EXCEL = "face_landmarks.xlsx"
OUTPUT_IMG_DIR = "sw_proj/face_with_points"

# =========================
# CLASS DEFINITION
# =========================
class FaceLandmarkProcessor:
    def __init__(self, predictor_path):
        self.detector = dlib.get_frontal_face_detector()  # Detector for faces
        self.predictor = dlib.shape_predictor(predictor_path)  # Landmark predictor
        self.landmark_data = []  # Store landmarks data for each image
        self.ratios = []  # Store ratio (R) for each image

        # Check if output image directory exists, if not, create it
        if not os.path.exists(OUTPUT_IMG_DIR):
            os.makedirs(OUTPUT_IMG_DIR)

    def calculate_distance(self, p1, p2):
        """Calculate Euclidean distance between two points (x1, y1) and (x2, y2)."""
        return np.linalg.norm(np.array(p1) - np.array(p2))

    def calculate_r_value(self, landmarks):
        """
        Calculate the face aspect ratio (R) = height / width
        Height is calculated between points 27 (nose tip) and 8 (chin)
        Width is calculated between points 0 (left jaw) and 16 (right jaw)
        """
        # Step 1: Calculate the height (distance between p27 and p8)
        height = self.calculate_distance(landmarks[27], landmarks[8])

        # Step 2: Calculate the width (distance between p0 and p16)
        width = self.calculate_distance(landmarks[0], landmarks[16])

        # Step 3: Return the ratio R = height / width
        return height / width if width != 0 else 0

    def draw_landmarks(self, image, landmarks, filename):
        """Draw landmarks and red lines (height and width) on the image."""
        for point in landmarks:
            cv2.circle(image, point, 2, (0, 255, 0), -1)  # Draw green circles on landmarks

        # Step 4: Draw red lines for face height and width (p27-p8, p0-p16)
        cv2.line(image, landmarks[27], landmarks[8], (0, 0, 255), 2)  # Red line for height
        cv2.line(image, landmarks[0], landmarks[16], (0, 0, 255), 2)  # Red line for width

        # Save image with landmarks
        output_path = os.path.join(OUTPUT_IMG_DIR, filename)
        cv2.imwrite(output_path, image)
        print(f"Saved image with landmarks: {output_path}")

    def process_images(self, folder):
        """
        Process all images in the provided folder, detecting faces, calculating landmarks,
        and drawing landmarks on images.
        """
        for filename in os.listdir(folder):
            if filename.lower().endswith((".jpg", ".jpeg", ".png")):
                image_path = os.path.join(folder, filename)
                print(f"Processing {image_path}...")

                image = cv2.imread(image_path)
                gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
                faces = self.detector(gray)  # Detect faces in the image

                # Step 5: Process each face detected
                for face in faces:
                    landmarks_obj = self.predictor(gray, face)
                    landmarks = [(landmarks_obj.part(i).x, landmarks_obj.part(i).y) for i in range(68)]

                    # Draw landmarks and save image with points
                    self.draw_landmarks(image.copy(), landmarks, filename)

                    # Step 6: Collect landmark data in a dictionary for saving to Excel
                    row = {"image": filename}
                    for i, (x, y) in enumerate(landmarks):
                        row[f"point_{i}_x"] = x
                        row[f"point_{i}_y"] = y
                    self.landmark_data.append(row)

                    # Step 7: Calculate R (aspect ratio) for each image and store it
                    r_value = self.calculate_r_value(landmarks)
                    self.ratios.append({"image": filename, "R": r_value})

    def save_to_excel(self):
        """Save the landmarks data, R values, and statistics to an Excel file."""
        writer = pd.ExcelWriter(OUTPUT_EXCEL, engine='openpyxl')

        # Step 8: Save landmarks data to Excel sheet
        df_landmarks = pd.DataFrame(self.landmark_data)
        df_landmarks.to_excel(writer, sheet_name="Overview", index=False)

        # Table formatting for the Overview sheet (optional based on TABLE_EXCEL flag)
        if TABLE_EXCEL:
            wb = writer.book
            ws = wb["Overview"]
            table = Table(displayName="FaceLandmarkTable", ref=f"A1:{get_column_letter(len(df_landmarks.columns))}{len(df_landmarks)+1}")
            style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
            table.tableStyleInfo = style
            ws.add_table(table)

        # Step 9: Save R values to another Excel sheet
        df_r = pd.DataFrame(self.ratios)
        df_r.to_excel(writer, sheet_name="R_values", index=False)

        # Step 10: Calculate statistics (mean, standard deviation, variance) for R values
        mean_r = df_r["R"].mean()  # Mean of R
        std_r = df_r["R"].std()  # Standard deviation of R
        variance = ((df_r["R"] - mean_r) ** 2).sum() / (len(df_r) - 1)  # Variance

        # Step 11: Save statistics to a new sheet in the Excel file
        stats_df = pd.DataFrame({
            "Mean R": [mean_r],
            "Std Dev": [std_r],
            "Variance": [variance]
        })
        stats_df.to_excel(writer, sheet_name="Stats", index=False)

        # Step 12: Create a histogram of the R values (Gaussian distribution plot)
        plt.figure(figsize=(6, 4))
        plt.hist(df_r["R"], bins=10, color='skyblue', edgecolor='black', density=True)
        plt.title("Distribution of R values")
        plt.xlabel("R")
        plt.ylabel("Frequency")
        plt.grid(True)
        plt.tight_layout()
        plot_path = "r_distribution.png"
        plt.savefig(plot_path)
        plt.close()

        # Step 13: Add the plot as an image to a new sheet in the Excel file
        wb = writer.book
        ws = wb.create_sheet("Distribution")
        img = ExcelImage(plot_path)
        img.anchor = 'A1'
        ws.add_image(img)

        writer.close()
        print(f"All data saved to {OUTPUT_EXCEL}")

# =========================
# MAIN EXECUTION
# =========================
if __name__ == "__main__":
    predictor_path = "sw_proj/shape_predictor_68_face_landmarks.dat"
    processor = FaceLandmarkProcessor(predictor_path)
    processor.process_images(IMAGE_FOLDER)
    processor.save_to_excel()
